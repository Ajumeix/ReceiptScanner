import sys
import cv2
import re
import json
import os
import pytesseract
from PIL import Image
from datetime import date
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout,
    QVBoxLayout, QPushButton, QLabel, QFileDialog,
    QScrollArea, QListWidget, QListWidgetItem, QStatusBar,
    QLineEdit, QComboBox, QFormLayout, QGroupBox,
    QDialog, QDialogButtonBox, QInputDialog, QMessageBox,
    QProgressBar, QSpinBox, QCheckBox, QTabWidget
)
from PyQt5.QtGui import QPixmap, QPalette, QColor, QFont, QPainter, QPen
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QRect, QPoint

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

SETTINGS_FILE = "settings.json"

DEFAULT_SETTINGS = {
    "dark_mode": False,
    "save_directory": "",
    "filename_template": "receipts_{date}_batch{n}",
    "particulars": ["Food", "Snack", "Transport", "Supplies", "Entertainment", "Other"],
    "batch_warn": 5
}

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, 'r') as f:
            s = json.load(f)
            for k, v in DEFAULT_SETTINGS.items():
                if k not in s:
                    s[k] = v
            return s
    return dict(DEFAULT_SETTINGS)

def save_settings(settings):
    with open(SETTINGS_FILE, 'w') as f:
        json.dump(settings, f, indent=2)

def preprocess(path):
    img = cv2.imread(path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, h=10)
    thresh = cv2.adaptiveThreshold(
        denoised, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY, 31, 15
    )
    return thresh

def preprocess_crop(img, x, y, w, h):
    crop = img[y:y+h, x:x+w]
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    gray = cv2.resize(gray, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)
    denoised = cv2.fastNlMeansDenoising(gray, h=10)
    _, thresh = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return thresh

def extract_vat_tin(text):
    match = re.search(r'(?:VAT\s*REG\s*)?TIN[:\s#]*([0-9]{9,15}[-\s]?[0-9]*)', text, re.IGNORECASE)
    return match.group(1).strip() if match else ""

def extract_total(text):
    matches = re.findall(r'(?:eat.?in\s*total|total)[^\d]*([0-9,]+\.[0-9]{2})', text, re.IGNORECASE)
    return matches[-1].strip() if matches else ""

def extract_date(text):
    match = re.search(r'(\d{2}/\d{2}/\d{4})', text)
    return match.group(1) if match else date.today().strftime("%m/%d/%Y")

def apply_dark_mode(app, enabled):
    if enabled:
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(30, 30, 30))
        palette.setColor(QPalette.WindowText, QColor(220, 220, 220))
        palette.setColor(QPalette.Base, QColor(45, 45, 45))
        palette.setColor(QPalette.AlternateBase, QColor(35, 35, 35))
        palette.setColor(QPalette.ToolTipBase, QColor(220, 220, 220))
        palette.setColor(QPalette.ToolTipText, QColor(220, 220, 220))
        palette.setColor(QPalette.Text, QColor(220, 220, 220))
        palette.setColor(QPalette.Button, QColor(50, 50, 50))
        palette.setColor(QPalette.ButtonText, QColor(220, 220, 220))
        palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        palette.setColor(QPalette.HighlightedText, QColor(255, 255, 255))
        app.setPalette(palette)
    else:
        app.setPalette(app.style().standardPalette())

class ScanThread(QThread):
    done = pyqtSignal(str, str)

    def __init__(self, path):
        super().__init__()
        self.path = path

    def run(self):
        processed = preprocess(self.path)
        pil_img = Image.fromarray(processed)
        raw = pytesseract.image_to_string(pil_img, config='--psm 6 --oem 3')
        self.done.emit(raw, self.path)

class ImageCanvas(QLabel):
    region_selected = pyqtSignal(int, int, int, int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.selection_mode = False
        self.drawing = False
        self.start_point = QPoint()
        self.current_rect = QRect()
        self.scale = 1.0
        self.setMouseTracking(True)

    def enable_selection(self):
        self.selection_mode = True
        self.setCursor(Qt.CrossCursor)

    def disable_selection(self):
        self.selection_mode = False
        self.drawing = False
        self.current_rect = QRect()
        self.setCursor(Qt.ArrowCursor)
        self.update()

    def mousePressEvent(self, event):
        if self.selection_mode and event.button() == Qt.LeftButton:
            self.drawing = True
            self.start_point = event.pos()
            self.current_rect = QRect(self.start_point, self.start_point)

    def mouseMoveEvent(self, event):
        if self.drawing:
            self.current_rect = QRect(self.start_point, event.pos()).normalized()
            self.update()

    def mouseReleaseEvent(self, event):
        if self.drawing and self.selection_mode:
            self.drawing = False
            rect = QRect(self.start_point, event.pos()).normalized()
            if rect.width() > 5 and rect.height() > 5:
                x = int(rect.x() / self.scale)
                y = int(rect.y() / self.scale)
                w = int(rect.width() / self.scale)
                h = int(rect.height() / self.scale)
                self.region_selected.emit(x, y, w, h)
            self.disable_selection()

    def paintEvent(self, event):
        super().paintEvent(event)
        if self.drawing and not self.current_rect.isNull():
            painter = QPainter(self)
            pen = QPen(QColor(255, 50, 50), 2, Qt.DashLine)
            painter.setPen(pen)
            fill = QColor(255, 50, 50, 40)
            painter.fillRect(self.current_rect, fill)
            painter.drawRect(self.current_rect)

    def load_image(self, path):
        pixmap = QPixmap(path)
        max_w, max_h = 360, 680
        self.scale = min(max_w / pixmap.width(), max_h / pixmap.height(), 1.0)
        scaled = pixmap.scaled(
            int(pixmap.width() * self.scale),
            int(pixmap.height() * self.scale),
            Qt.KeepAspectRatio, Qt.SmoothTransformation
        )
        self.setPixmap(scaled)

class SettingsDialog(QDialog):
    def __init__(self, settings, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Settings")
        self.setMinimumWidth(420)
        self.settings = dict(settings)

        layout = QVBoxLayout(self)
        tabs = QTabWidget()
        layout.addWidget(tabs)

        general = QWidget()
        gen_layout = QFormLayout(general)

        self.dark_toggle = QCheckBox("Enable Dark Mode")
        self.dark_toggle.setChecked(self.settings.get("dark_mode", False))
        gen_layout.addRow(self.dark_toggle)

        self.batch_warn_spin = QSpinBox()
        self.batch_warn_spin.setRange(1, 50)
        self.batch_warn_spin.setValue(self.settings.get("batch_warn", 5))
        gen_layout.addRow("Batch warning at:", self.batch_warn_spin)

        self.filename_input = QLineEdit(self.settings.get("filename_template", DEFAULT_SETTINGS["filename_template"]))
        gen_layout.addRow("Filename template:", self.filename_input)
        gen_layout.addRow(QLabel("  Use {date} and {n} as placeholders"))

        dir_layout = QHBoxLayout()
        self.dir_input = QLineEdit(self.settings.get("save_directory", ""))
        self.dir_input.setPlaceholderText("Default: ask every time")
        btn_browse = QPushButton("Browse")
        btn_browse.clicked.connect(self.browse_dir)
        dir_layout.addWidget(self.dir_input)
        dir_layout.addWidget(btn_browse)
        gen_layout.addRow("Save directory:", dir_layout)

        tabs.addTab(general, "General")

        particulars_tab = QWidget()
        part_layout = QVBoxLayout(particulars_tab)
        part_layout.addWidget(QLabel("Particulars dropdown options:"))

        self.part_list = QListWidget()
        for p in self.settings.get("particulars", []):
            self.part_list.addItem(p)
        part_layout.addWidget(self.part_list)

        part_btn_layout = QHBoxLayout()
        btn_add = QPushButton("Add")
        btn_add.clicked.connect(self.add_particular)
        btn_remove = QPushButton("Remove Selected")
        btn_remove.clicked.connect(self.remove_particular)
        btn_up = QPushButton("Move Up")
        btn_up.clicked.connect(self.move_up)
        btn_down = QPushButton("Move Down")
        btn_down.clicked.connect(self.move_down)
        part_btn_layout.addWidget(btn_add)
        part_btn_layout.addWidget(btn_remove)
        part_btn_layout.addWidget(btn_up)
        part_btn_layout.addWidget(btn_down)
        part_layout.addLayout(part_btn_layout)

        tabs.addTab(particulars_tab, "Particulars")

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def browse_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Select Save Directory")
        if d:
            self.dir_input.setText(d)

    def add_particular(self):
        text, ok = QInputDialog.getText(self, "Add Particular", "Enter name:")
        if ok and text.strip():
            self.part_list.addItem(text.strip())

    def remove_particular(self):
        row = self.part_list.currentRow()
        if row >= 0:
            self.part_list.takeItem(row)

    def move_up(self):
        row = self.part_list.currentRow()
        if row > 0:
            item = self.part_list.takeItem(row)
            self.part_list.insertItem(row - 1, item)
            self.part_list.setCurrentRow(row - 1)

    def move_down(self):
        row = self.part_list.currentRow()
        if row < self.part_list.count() - 1:
            item = self.part_list.takeItem(row)
            self.part_list.insertItem(row + 1, item)
            self.part_list.setCurrentRow(row + 1)

    def get_settings(self):
        self.settings["dark_mode"] = self.dark_toggle.isChecked()
        self.settings["batch_warn"] = self.batch_warn_spin.value()
        self.settings["filename_template"] = self.filename_input.text().strip()
        self.settings["save_directory"] = self.dir_input.text().strip()
        self.settings["particulars"] = [
            self.part_list.item(i).text()
            for i in range(self.part_list.count())
        ]
        return self.settings

class ReceiptScanner(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Receipt Scanner")
        self.setMinimumSize(1100, 750)
        self.settings = load_settings()
        self.batch = []
        self.batch_images = []
        self.image_paths = []
        self.current_index = 0
        self.current_image_path = None
        self.pick_target = None
        self.batch_counter = 1

        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(8)
        main_layout.setContentsMargins(12, 12, 12, 12)

        toolbar = QHBoxLayout()
        btn_import = QPushButton("Import Receipts")
        btn_import.setFixedHeight(36)
        btn_import.clicked.connect(self.import_images)

        btn_settings = QPushButton("Settings")
        btn_settings.setFixedHeight(36)
        btn_settings.clicked.connect(self.open_settings)

        self.filename_input = QLineEdit()
        self.filename_input.setFixedHeight(36)
        self.filename_input.setText(self.generate_filename())

        btn_export = QPushButton("Export to Excel")
        btn_export.setFixedHeight(36)
        btn_export.clicked.connect(self.export_excel)
        btn_export.setStyleSheet("background-color: #2a82da; color: white; font-weight: bold;")

        toolbar.addWidget(btn_import)
        toolbar.addWidget(btn_settings)
        toolbar.addStretch()
        toolbar.addWidget(QLabel("Filename:"))
        toolbar.addWidget(self.filename_input)
        toolbar.addWidget(btn_export)
        main_layout.addLayout(toolbar)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setRange(0, 0)
        self.progress.setFixedHeight(6)
        main_layout.addWidget(self.progress)

        content = QHBoxLayout()
        content.setSpacing(12)
        main_layout.addLayout(content)

        left = QVBoxLayout()
        content.addLayout(left)

        nav = QHBoxLayout()
        self.btn_prev = QPushButton("◀ Prev")
        self.btn_prev.setFixedHeight(30)
        self.btn_prev.clicked.connect(self.prev_image)
        self.btn_prev.setEnabled(False)

        self.image_counter = QLabel("No images")
        self.image_counter.setAlignment(Qt.AlignCenter)

        self.btn_next = QPushButton("Next ▶")
        self.btn_next.setFixedHeight(30)
        self.btn_next.clicked.connect(self.next_image)
        self.btn_next.setEnabled(False)

        btn_clear_images = QPushButton("Clear Images")
        btn_clear_images.setFixedHeight(30)
        btn_clear_images.clicked.connect(self.clear_images)

        nav.addWidget(self.btn_prev)
        nav.addWidget(self.image_counter)
        nav.addWidget(self.btn_next)
        nav.addWidget(btn_clear_images)
        left.addLayout(nav)

        self.canvas = ImageCanvas()
        self.canvas.setText("No image loaded")
        self.canvas.setAlignment(Qt.AlignTop | Qt.AlignHCenter)
        self.canvas.setStyleSheet("border: 1px solid #ccc; border-radius: 6px;")
        self.canvas.region_selected.connect(self.on_region_selected)

        scroll_left = QScrollArea()
        scroll_left.setWidget(self.canvas)
        scroll_left.setWidgetResizable(True)
        scroll_left.setMinimumWidth(380)
        left.addWidget(scroll_left)

        right = QVBoxLayout()
        right.setSpacing(8)
        content.addLayout(right)

        form_group = QGroupBox("Receipt Details")
        form = QFormLayout(form_group)
        form.setSpacing(10)

        self.date_input = QLineEdit()
        self.date_input.setPlaceholderText("MM/DD/YYYY")
        self.date_input.setFixedHeight(32)
        form.addRow("Date:", self.date_input)

        self.particulars_combo = QComboBox()
        self.particulars_combo.setFixedHeight(32)
        self.particulars_combo.addItems(self.settings["particulars"])
        form.addRow("Particulars:", self.particulars_combo)

        vat_layout = QHBoxLayout()
        self.vat_input = QLineEdit()
        self.vat_input.setPlaceholderText("Auto-filled from scan")
        self.vat_input.setFixedHeight(32)
        btn_pick_vat = QPushButton("Pick Area")
        btn_pick_vat.setFixedHeight(32)
        btn_pick_vat.setFixedWidth(80)
        btn_pick_vat.setStyleSheet("background-color: #e67e22; color: white;")
        btn_pick_vat.clicked.connect(lambda: self.start_pick("vat"))
        vat_layout.addWidget(self.vat_input)
        vat_layout.addWidget(btn_pick_vat)
        form.addRow("VAT REG TIN:", vat_layout)

        self.shared_combo = QComboBox()
        self.shared_combo.setFixedHeight(32)
        self.shared_combo.addItems(["Solo", "Shared"])
        form.addRow("Solo / Shared:", self.shared_combo)

        total_layout = QHBoxLayout()
        self.total_input = QLineEdit()
        self.total_input.setPlaceholderText("Auto-filled from scan")
        self.total_input.setFixedHeight(32)
        btn_pick_total = QPushButton("Pick Area")
        btn_pick_total.setFixedHeight(32)
        btn_pick_total.setFixedWidth(80)
        btn_pick_total.setStyleSheet("background-color: #e67e22; color: white;")
        btn_pick_total.clicked.connect(lambda: self.start_pick("total"))
        total_layout.addWidget(self.total_input)
        total_layout.addWidget(btn_pick_total)
        form.addRow("Total:", total_layout)

        right.addWidget(form_group)

        btn_add = QPushButton("Add to Batch  [Enter]")
        btn_add.setFixedHeight(38)
        btn_add.setShortcut("Return")
        btn_add.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold;")
        btn_add.clicked.connect(self.add_to_batch)
        right.addWidget(btn_add)

        batch_header = QHBoxLayout()
        self.batch_label = QLabel("Batch: 0 receipts")
        self.batch_label.setFont(QFont("Arial", 10, QFont.Bold))
        btn_clear = QPushButton("Clear Batch")
        btn_clear.setFixedHeight(28)
        btn_clear.clicked.connect(self.clear_batch)
        batch_header.addWidget(self.batch_label)
        batch_header.addStretch()
        batch_header.addWidget(btn_clear)
        right.addLayout(batch_header)

        self.batch_list = QListWidget()
        self.batch_list.setAlternatingRowColors(True)
        right.addWidget(self.batch_list)

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage("Ready — import receipt images")

        apply_dark_mode(QApplication.instance(), self.settings.get("dark_mode", False))

    def generate_filename(self):
        template = self.settings.get("filename_template", DEFAULT_SETTINGS["filename_template"])
        today = date.today().strftime("%Y-%m-%d")
        return template.replace("{date}", today).replace("{n}", str(self.batch_counter))

    def import_images(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Select Receipts", "", "Images (*.png *.jpg *.jpeg)")
        if not paths:
            return
        self.image_paths = paths
        self.current_index = 0
        self.load_image(self.current_index)
        self.update_nav()

    def load_image(self, index):
        path = self.image_paths[index]
        self.current_image_path = path
        self.canvas.load_image(path)
        self.vat_input.clear()
        self.total_input.clear()
        self.date_input.clear()
        self.progress.setVisible(True)
        self.status.showMessage(f"Scanning image {index + 1} of {len(self.image_paths)}...")

        self.scan_thread = ScanThread(path)
        self.scan_thread.done.connect(self.on_scan_done)
        self.scan_thread.start()

    def update_nav(self):
        total = len(self.image_paths)
        self.image_counter.setText(f"{self.current_index + 1} / {total}")
        self.btn_prev.setEnabled(self.current_index > 0)
        self.btn_next.setEnabled(self.current_index < total - 1)

    def prev_image(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.load_image(self.current_index)
            self.update_nav()

    def next_image(self):
        if self.current_index < len(self.image_paths) - 1:
            self.current_index += 1
            self.load_image(self.current_index)
            self.update_nav()

    def clear_images(self):
        self.image_paths = []
        self.current_index = 0
        self.current_image_path = None
        self.canvas.clear()
        self.canvas.setText("No image loaded")
        self.image_counter.setText("No images")
        self.btn_prev.setEnabled(False)
        self.btn_next.setEnabled(False)
        self.vat_input.clear()
        self.total_input.clear()
        self.date_input.clear()
        self.status.showMessage("Images cleared")

    def on_scan_done(self, raw_text, path):
        self.progress.setVisible(False)
        self.vat_input.setText(extract_vat_tin(raw_text))
        self.total_input.setText(extract_total(raw_text))
        self.date_input.setText(extract_date(raw_text))
        self.status.showMessage("Scan done — review fields or use Pick Area if needed")

    def start_pick(self, target):
        if not self.current_image_path:
            QMessageBox.warning(self, "No Image", "Please import a receipt image first.")
            return
        self.pick_target = target
        self.canvas.enable_selection()
        self.status.showMessage(f"Draw a box around the {target.upper()} area on the image...")

    def on_region_selected(self, x, y, w, h):
        if not self.current_image_path or not self.pick_target:
            return

        img = cv2.imread(self.current_image_path)
        h_img, w_img = img.shape[:2]
        x = max(0, min(x, w_img - 1))
        y = max(0, min(y, h_img - 1))
        w = min(w, w_img - x)
        h = min(h, h_img - y)

        processed = preprocess_crop(img, x, y, w, h)
        pil_img = Image.fromarray(processed)

        configs = ['--psm 7 --oem 3', '--psm 6 --oem 3', '--psm 8 --oem 3']
        best = ""
        best_len = 0
        for config in configs:
            text = pytesseract.image_to_string(pil_img, config=config).strip()
            if len(text) > best_len:
                best_len = len(text)
                best = text

        best = re.sub(r'[^\w\s.,\-/:]', '', best).strip()

        if self.pick_target == "vat":
            nums = re.findall(r'[0-9]{3,}[-\s]?[0-9]*', best)
            self.vat_input.setText("-".join(nums) if nums else best)
            self.status.showMessage(f"VAT area scanned: {self.vat_input.text()}")
        elif self.pick_target == "total":
            nums = re.findall(r'[0-9,]+\.[0-9]{2}', best)
            self.total_input.setText(nums[0] if nums else best)
            self.status.showMessage(f"Total area scanned: {self.total_input.text()}")

        self.pick_target = None

    def add_to_batch(self):
        if not self.current_image_path:
            QMessageBox.warning(self, "No Image", "Please import receipt images first.")
            return
        if not self.date_input.text().strip():
            QMessageBox.warning(self, "Missing Date", "Please enter a date.")
            return

        entry = {
            "Date": self.date_input.text().strip(),
            "Particulars": self.particulars_combo.currentText(),
            "VAT REG TIN": self.vat_input.text().strip(),
            "Solo / Shared Receipt": self.shared_combo.currentText(),
            "Total": self.total_input.text().strip(),
        }
        self.batch.append(entry)
        self.batch_images.append(self.current_image_path)

        summary = f"{len(self.batch)}. {entry['Date']} | {entry['Particulars']} | {entry['Solo / Shared Receipt']} | ₱{entry['Total']}"
        self.batch_list.addItem(summary)
        self.batch_label.setText(f"Batch: {len(self.batch)} receipt(s)")

        warn = self.settings.get("batch_warn", 5)
        if len(self.batch) == warn:
            QMessageBox.information(self, "Batch Notice", f"You have {warn} receipts — consider exporting soon.")

        if self.current_index < len(self.image_paths) - 1:
            self.current_index += 1
            self.load_image(self.current_index)
            self.update_nav()
        else:
            self.vat_input.clear()
            self.total_input.clear()
            self.date_input.clear()
            self.status.showMessage(f"All images processed — {len(self.batch)} in batch")

    def clear_batch(self):
        if self.batch and QMessageBox.question(self, "Clear Batch", "Clear the batch?") != QMessageBox.Yes:
            return
        self.batch = []
        self.batch_images = []
        self.batch_list.clear()
        self.batch_label.setText("Batch: 0 receipts")
        self.batch_counter += 1
        self.filename_input.setText(self.generate_filename())
        self.status.showMessage("Batch cleared")

    def open_settings(self):
        dialog = SettingsDialog(self.settings, self)
        if dialog.exec_():
            self.settings = dialog.get_settings()
            save_settings(self.settings)
            self.particulars_combo.clear()
            self.particulars_combo.addItems(self.settings["particulars"])
            self.filename_input.setText(self.generate_filename())
            apply_dark_mode(QApplication.instance(), self.settings.get("dark_mode", False))
            self.status.showMessage("Settings saved")

    def export_excel(self):
        if not self.batch:
            QMessageBox.warning(self, "Empty Batch", "No receipts in batch to export.")
            return

        filename = self.filename_input.text().strip() or self.generate_filename()
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        save_dir = self.settings.get("save_directory", "")
        default_path = os.path.join(save_dir, filename) if save_dir and os.path.isdir(save_dir) else filename

        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", default_path, "Excel Files (*.xlsx)")
        if not path:
            return

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Receipt Data"
        headers = ["Date", "Particulars", "VAT REG TIN", "Solo / Shared Receipt", "Total"]
        ws1.append(headers)

        for col in range(1, 6):
            ws1.cell(row=1, column=col).font = Font(bold=True)

        for entry in self.batch:
            ws1.append([entry[h] for h in headers])

        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = max_len + 4

        ws2 = wb.create_sheet(title="Receipt Images")
        ws2.append(["#", "File", "Image"])
        for i, img_path in enumerate(self.batch_images):
            row = i + 2
            ws2.cell(row=row, column=1, value=i + 1)
            ws2.cell(row=row, column=2, value=os.path.basename(img_path))
            try:
                xl_img = XLImage(img_path)
                xl_img.width = 280
                xl_img.height = 380
                ws2.row_dimensions[row].height = 285
                ws2.add_image(xl_img, f"C{row}")
            except Exception as e:
                ws2.cell(row=row, column=3, value=f"Error: {e}")

        wb.save(path)
        self.batch_counter += 1
        self.filename_input.setText(self.generate_filename())
        self.status.showMessage(f"Exported {len(self.batch)} receipts to {path}")
        QMessageBox.information(self, "Export Done", f"Saved {len(self.batch)} receipts to:\n{path}")

app = QApplication(sys.argv)
window = ReceiptScanner()
window.show()
sys.exit(app.exec_())
